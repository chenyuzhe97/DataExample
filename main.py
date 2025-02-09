import pandas as pd
import re
from fuzzywuzzy import fuzz
from collections import Counter, defaultdict
import matplotlib.pyplot as plt
import os
# 额外引入 openpyxl 的填充工具
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog


class ReCreator:
    # 地区括号提取
    def parenthesis_extraction(self, text):
        match = re.search(r'\((.*?)\)', text)
        if match:
            return match.group(1)
        else:
            raise '未找到匹配项！'

    # 地区清洗
    def clean_location(self, location: str) -> str:
        import re
        location = re.sub(r'[()（）\-—]', ' ', location)  # 将括号等字符替换为空格
        location = re.sub(r'(省|市|区|县|自治区|自治州|盟|地区)$', '', location)  # 去掉后缀
        location = re.sub(r'\s+', ' ', location).strip()  # 去除多余空格
        return location


class PieChart:
    def __init__(self):
        self.pie_list = {}
        similarity_counter = Counter()
        weak1 = Counter()
        weak2 = Counter()
        self.pie_list["similarity"] = similarity_counter
        self.pie_list["weak1"] = weak1
        self.pie_list["weak2"] = weak2

    # 画柱状图
    def plot_bar_chart(self, name):
        counter = self.pie_list[name]
        labels = list(counter.keys())
        sizes = list(counter.values())

        plt.figure(figsize=(8, 6))
        plt.barh(labels, sizes, color='skyblue')
        plt.xlabel('Count')
        plt.ylabel('Categories')
        plt.title(name)

        for index, value in enumerate(sizes):
            plt.text(value, index, str(value), va='center')

        plt.show()

    def draw_all_bar_charts(self):
        for key in self.pie_list.keys():
            self.plot_bar_chart(key)

    # 画饼图
    def plot_pie_chart(self, name):
        counter = self.pie_list[name]
        labels = list(counter.keys())
        sizes = list(counter.values())

        plt.figure(figsize=(6, 6))
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.axis('equal')
        plt.title(name)
        plt.show()

    def draw_all_pie_chart(self):
        for key in self.pie_list.items():
            self.plot_pie_chart(key)

    def append(self, name, counter):
        self.pie_list[name] = counter

    def get(self, name):
        return self.pie_list[name]


class QuestionnaireChecker:
    def __init__(self, file_path):
        self.re_creator = ReCreator()
        self.pie_chart = PieChart()

        self.file_path = file_path
        self.df = pd.read_excel(file_path)
        self.file_path = os.path.splitext(file_path)[0]
        # 去掉所有空格（列名左右空格）
        self.df.columns = self.df.columns.str.strip()
        self.errors = []

        # 调试模式
        self.debug = False

        # 警告数量阈值（>= 此数量的警告，整行视为“严重”）
        self.WARNING_THRESHOLD = 3

    def show_columns(self):
        print("列名：", self.df.columns)

    def check_time(self):
        """ 检查问卷的回答时间是否合理 """
        for index, row in self.df.iterrows():
            # 假设时间格式是类似 "120s" 或 "75s"，先去掉最后的 "s"
            answer_time = int(row["所用时间"][:-1])
            if answer_time < 60:
                self.errors.append({
                    "行号": index + 1,  # 注意，这里 +1 仅表示数据的行号（不含表头）
                    "列名": "回答时间",
                    "错误类型": "严重",
                    "描述": "回答时间过短"
                })

    def check_ip_and_location(self):
        """ 检查IP与居住地是否匹配 """
        for index, row in self.df.iterrows():
            ip = row["来自IP"]
            location = row["4.目前居住省份城市"]
            if not self.is_ip_match_location(ip, location):
                self.errors.append({
                    "行号": index + 1,
                    "列名": "IP地址",
                    "错误类型": "警告",
                    "描述": f"IP与居住地不匹配{ip}与{location}"
                })
        if self.debug:
            self.pie_chart.plot_bar_chart('similarity')

    def check_age_consistency(self):
        """ 检查年龄前后是否一致 """
        for index, row in self.df.iterrows():
            if row["2.年龄:"] != row["13.年龄:"]:
                self.errors.append({
                    "行号": index + 1,
                    "列名": "年龄",
                    "错误类型": "严重",
                    "描述": "前后年龄不一致"
                })

    def check_working_consistency(self):
        """ 检查工作前后是否一致 """
        for index, row in self.df.iterrows():
            if row["3.职业:"] != row["14.职业:"]:
                self.errors.append({
                    "行号": index + 1,
                    "列名": "工作",
                    "错误类型": "严重",
                    "描述": "前后工作不一致"
                })

    def check_consumption_level(self):
        for index, row in self.df.iterrows():
            sport_product = row['17.您之前买过其他的智能健身类产品是什么?（多选）']
            sport_spend = row['18.您在其他智能健身类产品上投入了多少资金?']
            flag = self.is_product_match_spend(sport_product, sport_spend)
            if flag == 'warning':
                self.errors.append({
                    "行号": index + 1,
                    "列名": "工作",
                    "错误类型": "警告",
                    "描述": f"前后投入资金不一致,{sport_product},{sport_spend}"
                })
            elif flag == 'wrong':
                self.errors.append({
                    "行号": index + 1,
                    "列名": "工作",
                    "错误类型": "严重",
                    "描述": f"乱填,{sport_product},{sport_spend}"
                })

    def check_weak_relevance(self):
        weak1 = self.pie_chart.get('weak1')
        weak2 = self.pie_chart.get('weak2')
        for index, row in self.df.iterrows():
            question1 = row[
                '19.在锻炼时，如果你发现智能运动设备给出的训练计划时间过长，可能会影响到你的工作计划和外出活动，你最有可能的反应是:']
            if question1 == 'D. 继续按计划锻炼，认为工作可以后延':
                self.errors.append({
                    "行号": index + 1,
                    "列名": "弱相关问题",
                    "错误类型": "警告",
                    "描述": f"人工复查该选项，不太不符合正常人案例：D. 继续按计划锻炼，认为工作可以后延"
                })
            question2 = row[
                '24  假设你正在进行一项高强度的力量训练，而智能运动健身足垫突然给出了一条建议，让你休息片刻。你最可能的反应是什么?']
            if ('A' in question1 or 'B' in question1 or 'C' in question1) and 'C' in question2:
                self.errors.append({
                    "行号": index + 1,
                    "列名": "弱相关问题",
                    "错误类型": "警告",
                    "描述": f"人工复查该选项，不太不符合正常人逻辑，前面听话了，后面又不听话"
                })
            weak1[question1] += 1
            weak2[question2] += 1

    def check_warnings(self):
        """ 统计警告类错误（示例） """
        pass
        # 若有需要，可自行在此处添加“警告”级别的逻辑

    def save_results(self):
        """
        生成错误报告并标注Excel。

        1）将 self.errors 写入错误报告 Sheet；
        2）在 "问卷数据" Sheet 中，对出现问题的整行进行填充：
           - "严重" → 红色
           - "警告" → 黄色
           - 如果同一行的警告数量 >= self.WARNING_THRESHOLD(默认3) → 红色
        """
        error_df = pd.DataFrame(self.errors)
        error_df.to_excel(self.file_path + "_errors.xlsx", index=False)

        # ============= 1. 统计每行错误情况 =============
        from collections import defaultdict
        row_warning_count = defaultdict(int)
        row_serious_count = defaultdict(int)

        for err in self.errors:
            r = err["行号"]
            if err["错误类型"] == "严重":
                row_serious_count[r] += 1
            elif err["错误类型"] == "警告":
                row_warning_count[r] += 1

        # ============= 2. 准备行着色信息 =============
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
        yellow_fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")  # 黄色

        row_colors = {}
        all_error_rows = set(row_warning_count.keys()) | set(row_serious_count.keys())
        for row_idx in all_error_rows:
            serious_count = row_serious_count[row_idx]
            warning_count = row_warning_count[row_idx]

            # 规则：如果同一行包含严重错误 或 警告数 ≥ WARNING_THRESHOLD，则红色
            if serious_count > 0 or warning_count >= self.WARNING_THRESHOLD:
                row_colors[row_idx] = red_fill
            else:
                # 说明该行没有严重错误，但有警告(且警告数 < 阈值)，则标黄
                row_colors[row_idx] = yellow_fill

        # ============= 3. 写入数据与错误报告 =============
        with pd.ExcelWriter(self.file_path + "_checked_questionnaire.xlsx",
                            engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="问卷数据", index=False)
            error_df.to_excel(writer, sheet_name="错误报告", index=False)

            # ============= 4. 对原数据进行行着色 =============
            workbook = writer.book
            worksheet = workbook["问卷数据"]

            max_col = worksheet.max_column
            # 因为 df.to_excel(index=False) 后，第1行为表头，第2行才是数据
            # self.errors 里的行号（index+1） 代表 Excel 数据行位置（不含表头）
            # 因此在给 Worksheet 着色时，需要对行号再 +1
            for row_idx, fill_color in row_colors.items():
                excel_row = row_idx + 1
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.fill = fill_color

    def is_product_match_spend(self, sport, spend):
        if sport == '没有买过（跳转第14题）' and spend == '1000元以下':
            return 'ok'
        elif sport == '智能健身手环/手表' and (spend == '1000元以下' or sport == '1000-5000元'):
            return 'ok'
        elif sport == '智能哑铃' and spend == '1000元以下':
            return 'ok'
        elif '智能健身手环/手表' in sport and '智能哑铃' in sport and spend == '1000元以下':
            return 'ok'
        elif "智能跑步机" in sport and spend != '1000元以下':
            return 'ok'
        elif "智能健身镜" in sport and spend != '1000元以下':
            return 'ok'
        elif "智能健身车" in sport and spend != '1000元以下':
            return 'ok'
        elif sport != '没有买过（跳转第14题）' and '没有买过（跳转第14题）' in sport:
            return 'wrong'
        else:
            return 'warning'

    def is_ip_match_location(self, ip, location):
        """ 利用模糊匹配判断 IP 所在地与填写的居住地是否相似 """
        ip = self.re_creator.parenthesis_extraction(ip)
        ip = self.re_creator.clean_location(ip)
        location = self.re_creator.clean_location(location)
        similarity = fuzz.ratio(ip, location)
        self.pie_chart.get('similarity')[similarity] += 1

        if self.debug:
            print(ip, location, f"相似度: {similarity}%")
        return similarity >= 40

    def run_checks(self):
        """ 按需运行所有检查 """
        self.check_time()
        self.check_ip_and_location()
        self.check_age_consistency()
        self.check_working_consistency()
        self.check_consumption_level()
        self.check_weak_relevance()
        self.check_warnings()
        self.save_results()


def select_and_process_files():
    file_paths = filedialog.askopenfilenames(
        title="请选择要处理的 Excel 文件（可多选）",
        filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
    )
    if not file_paths:
        print("未选择任何文件。")
        return

    for file_path in file_paths:
        print(f"开始处理: {file_path}")
        checker = QuestionnaireChecker(file_path)
        checker.run_checks()
        print(f"处理完成: {file_path}")
    print("所有文件处理完毕！")


if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()
    root.title("问卷检查工具")

    # 创建一个按钮，用于打开文件对话框
    btn = tk.Button(root, text="选择文件并处理", command=select_and_process_files)
    btn.pack(padx=20, pady=20)

    # 进入主事件循环
    root.mainloop()
